"""
Sonuçlar klasöründen (sonuclar/) Excel dosyalarını okuyup geçmiş listesi ve detay sağlar.
"""
import os
import json
import logging
from typing import List, Dict, Optional, Any
from datetime import datetime

logger = logging.getLogger(__name__)

SONUCLAR_DIR = "sonuclar"


def _parse_excel_run(file_path: str, file_id: str) -> Optional[Dict]:
    """
    Excel dosyasından özet, kesim planı ve rulo durumu çıkarır.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Excel okunamadı %s: %s", file_path, str(e))
        return None

    summary = {"totalCost": 0, "totalFire": 0, "totalStock": 0, "openedRolls": 0}
    cutting_plan: List[Dict] = []
    roll_status: List[Dict] = []

    try:
        if "Ozet" in wb.sheetnames:
            ws = wb["Ozet"]
            for row in range(4, 10):
                label = ws.cell(row=row, column=1).value
                val = ws.cell(row=row, column=2).value
                if label == "Toplam Maliyet" and val is not None:
                    summary["totalCost"] = float(val) if isinstance(val, (int, float)) else float(str(val).replace(",", "."))
                elif label and "Fire" in str(label) and val is not None:
                    summary["totalFire"] = float(val) if isinstance(val, (int, float)) else float(str(val).replace(",", "."))
                elif label and "Stok" in str(label) and val is not None:
                    summary["totalStock"] = float(val) if isinstance(val, (int, float)) else float(str(val).replace(",", "."))
                elif label and "Rulo" in str(label) and val is not None:
                    summary["openedRolls"] = int(val) if isinstance(val, (int, float)) else int(float(str(val)))

        if "Kesim_Plani" in wb.sheetnames:
            ws = wb["Kesim_Plani"]
            for row in range(4, ws.max_row + 1):
                siparis_cell = ws.cell(row=row, column=1).value
                if not siparis_cell or "Sipariş" not in str(siparis_cell):
                    continue
                try:
                    order_id = int(str(siparis_cell).replace("Sipariş", "").strip())
                except ValueError:
                    continue
                rulo_str = ws.cell(row=row, column=2).value or ""
                roll_id = int(str(rulo_str).replace("Rulo", "").strip()) if rulo_str else 0
                panel_count = _to_float(ws.cell(row=row, column=3).value) or 0
                panel_width = _to_float(ws.cell(row=row, column=4).value) or 0
                tonnage = _to_float(ws.cell(row=row, column=5).value) or 0
                m2 = _to_float(ws.cell(row=row, column=6).value) or 0
                cutting_plan.append({
                    "orderId": order_id,
                    "rollId": roll_id,
                    "panelCount": int(panel_count),
                    "panelWidth": panel_width,
                    "tonnage": tonnage,
                    "m2": m2,
                })

        if "Rulo_Durumu" in wb.sheetnames:
            ws = wb["Rulo_Durumu"]
            for row in range(4, ws.max_row + 1):
                rulo_str = ws.cell(row=row, column=1).value or ""
                if not rulo_str or "Rulo" not in str(rulo_str):
                    continue
                try:
                    roll_id = int(str(rulo_str).replace("Rulo", "").strip())
                except ValueError:
                    continue
                total_tonnage = _to_float(ws.cell(row=row, column=2).value) or 0
                used = _to_float(ws.cell(row=row, column=3).value) or 0
                remaining = _to_float(ws.cell(row=row, column=4).value) or 0
                stock = _to_float(ws.cell(row=row, column=5).value) or 0
                fire = _to_float(ws.cell(row=row, column=6).value) or 0
                orders_str = str(ws.cell(row=row, column=7).value or "0")
                orders_used = int(orders_str.split("/")[0]) if "/" in orders_str else int(float(orders_str) if orders_str else 0)
                roll_status.append({
                    "rollId": roll_id,
                    "totalTonnage": total_tonnage,
                    "used": used,
                    "remaining": remaining,
                    "fire": fire,
                    "stock": stock,
                    "ordersUsed": orders_used,
                })

        wb.close()
    except Exception as e:
        logger.exception("Excel parse hatası %s: %s", file_path, str(e))
        return None

    mtime = os.path.getmtime(file_path)
    created_at = datetime.fromtimestamp(mtime).isoformat()

    return {
        "id": file_id,
        "file_id": file_id,
        "created_at": created_at,
        "summary": summary,
        "cutting_plan": cutting_plan,
        "roll_status": roll_status,
        "status": "Optimal",
    }


def _to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").strip())
    except ValueError:
        return None


def list_runs_local(limit: int = 50, offset: int = 0) -> List[Dict]:
    """
    sonuclar/ klasöründeki Excel dosyalarından geçmiş listesini oluşturur.
    """
    if not os.path.exists(SONUCLAR_DIR):
        return []

    runs = []
    prefix = "cozum_raporu_"
    suffix = ".xlsx"
    files = [
        (name, os.path.getmtime(os.path.join(SONUCLAR_DIR, name)))
        for name in os.listdir(SONUCLAR_DIR)
        if name.startswith(prefix) and name.endswith(suffix)
    ]
    files.sort(key=lambda x: x[1], reverse=True)

    for name, _ in files:
        file_id = name[len(prefix) : -len(suffix)]
        file_path = os.path.join(SONUCLAR_DIR, name)
        if not os.path.isfile(file_path):
            continue
        data = _parse_excel_run(file_path, file_id)
        if data:
            runs.append({
                "id": data["file_id"],
                "file_id": data["file_id"],
                "created_at": data["created_at"],
                "summary": data["summary"],
                "status": data["status"],
            })

    return runs[offset : offset + limit]


def get_run_local(file_id: str) -> Optional[Dict]:
    """
    Belirli file_id için sonucu sonuclar/ klasöründen okur.
    """
    file_path = os.path.join(SONUCLAR_DIR, f"cozum_raporu_{file_id}.xlsx")
    if not os.path.exists(file_path):
        return None
    return _parse_excel_run(file_path, file_id)
