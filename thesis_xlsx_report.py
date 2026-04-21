"""
Ortak Excel rapor üreticisi.

main/sonuclar/cozum_raporu_*.xlsx formatına birebir uyan çok sayfalı (Kullanici_Veri_Girisi,
Ozet, Rulo_Durumu, Detayli_Kesim_Plani, Uretim_Adimlari) + opsiyonel Grafikler sayfası üretir.
Tez doğrulama, OFAT duyarlılık ve main batch akışlarında ortak şablon olarak kullanılır.
"""

from __future__ import annotations

import os
from typing import Any, Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# API / ön yüz ile aynı: sipariş başına max rulo için üst sınır yok anlamında gönderilen değer.
MAX_ROLLS_PER_ORDER_UNLIMITED_SENTINEL = 999999

BASLIK_DOLGU = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
ALT_BASLIK_DOLGU = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
BASLIK_FONT = Font(bold=True)
BUYUK_BASLIK = Font(bold=True, size=14)
ORTA_BASLIK = Font(bold=True, size=12)


def _yaz_etiket_deger(ws, row: int, etiket: str, deger: Any, *, kalın_etiket: bool = True) -> None:
    """
    Tek satıra etiket + değer yazar.

    Args:
        ws: openpyxl sayfası
        row: Satır numarası
        etiket: A sütununa yazılacak etiket
        deger: B sütununa yazılacak değer
        kalın_etiket: True ise etiket bold
    """
    c1 = ws.cell(row=row, column=1, value=etiket)
    if kalın_etiket:
        c1.font = BASLIK_FONT
    ws.cell(row=row, column=2, value=deger)


def _bolum_baslik(ws, row: int, baslik: str, merge_to: str = "B") -> None:
    """
    Bir bölüm başlığını gri zeminli hücre olarak yazar.

    Args:
        ws: openpyxl sayfası
        row: Satır numarası
        baslik: Başlık metni
        merge_to: Birleştirme sütun harfi (A'dan bu harfe)
    """
    cell = ws.cell(row=row, column=1, value=baslik)
    cell.font = ORTA_BASLIK
    cell.fill = BASLIK_DOLGU
    ws.merge_cells(f"A{row}:{merge_to}{row}")


def _tablo_yaz(
    ws,
    baslangic_satir: int,
    basliklar: Sequence[str],
    satirlar: Sequence[Sequence[Any]],
) -> int:
    """
    Başlık satırı + veri satırları olan sabit tablo yazar.

    Args:
        ws: openpyxl sayfası
        baslangic_satir: İlk satır (başlık için)
        basliklar: Kolon başlık listesi
        satirlar: Veri satırlarının iterable'ı

    Returns:
        Son yazılan satır numarası
    """
    for col_idx, baslik in enumerate(basliklar, start=1):
        c = ws.cell(row=baslangic_satir, column=col_idx, value=baslik)
        c.font = BASLIK_FONT
        c.fill = BASLIK_DOLGU
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r = baslangic_satir
    for row_data in satirlar:
        r += 1
        for col_idx, deger in enumerate(row_data, start=1):
            ws.cell(row=r, column=col_idx, value=deger)
    return r


def _sayfa_1_veri_girisi(wb: Workbook, meta: Dict[str, Any]) -> None:
    """
    Sayfa 1: Kullanıcı Veri Girişi. Senaryo, malzeme, tonaj, rulo tonajları,
    siparişler (m²) ve maliyet katsayılarını yazar.

    Args:
        wb: Workbook
        meta: scenario_meta sözlüğü (senaryo_adi, guvenlik_payi, kalinlik_mm, yogunluk_g_cm3,
              rulolar_ton, siparisler, maliyetler, toplam_talep_ton, toplam_rulo_kapasitesi_ton,
              max_siparis_per_rulo, max_rulo_per_siparis)
    """
    ws = wb.create_sheet("Kullanici_Veri_Girisi", 0)
    ws["A1"] = "KULLANICI VERİ GİRİŞİ"
    ws["A1"].font = BUYUK_BASLIK
    ws.merge_cells("A1:B1")

    row = 3
    _bolum_baslik(ws, row, "SENARYO BİLGİSİ")
    row += 1
    _yaz_etiket_deger(ws, row, "Seçilen Senaryo", meta.get("senaryo_adi", ""))
    row += 1
    gp = meta.get("guvenlik_payi")
    if gp is not None:
        _yaz_etiket_deger(ws, row, "Güvenlik Payı (%)", gp)
        row += 1
    aciklama = meta.get("aciklama")
    if aciklama:
        _yaz_etiket_deger(ws, row, "Açıklama", aciklama)
        row += 1
    row += 1

    _bolum_baslik(ws, row, "MALZEME ÖZELLİKLERİ")
    row += 1
    _yaz_etiket_deger(ws, row, "Kalınlık (mm)", meta.get("kalinlik_mm"))
    row += 1
    _yaz_etiket_deger(ws, row, "Yoğunluk (g/cm³)", meta.get("yogunluk_g_cm3"))
    row += 2

    _bolum_baslik(ws, row, "TONAJ BİLGİSİ")
    row += 1
    _yaz_etiket_deger(
        ws, row, "Toplam İhtiyaç (Ton)", round(float(meta.get("toplam_talep_ton") or 0), 4)
    )
    row += 1
    _yaz_etiket_deger(
        ws,
        row,
        "Toplam Rulo Kapasitesi (Ton)",
        round(float(meta.get("toplam_rulo_kapasitesi_ton") or 0), 2),
    )
    row += 1
    mspr = meta.get("max_siparis_per_rulo")
    if mspr is not None:
        _yaz_etiket_deger(
            ws,
            row,
            "Bir Rulo Maksimum Sipariş Sayısı",
            "Sınırsız" if int(mspr) >= 999 else mspr,
        )
        row += 1
    mrps = meta.get("max_rulo_per_siparis")
    if mrps is not None:
        mrps_int = int(mrps)
        mrps_goster = (
            "sonsuz"
            if mrps_int >= MAX_ROLLS_PER_ORDER_UNLIMITED_SENTINEL
            else mrps_int
        )
        _yaz_etiket_deger(ws, row, "Sipariş Başına Maksimum Rulo Sayısı", mrps_goster)
        row += 1
    row += 1

    _bolum_baslik(ws, row, "RULO TONAJLARI (Ton)")
    row += 1
    r1 = ws.cell(row=row, column=1, value="Rulo No")
    r1.font = BASLIK_FONT
    r1.fill = ALT_BASLIK_DOLGU
    r2 = ws.cell(row=row, column=2, value="Tonaj (Ton)")
    r2.font = BASLIK_FONT
    r2.fill = ALT_BASLIK_DOLGU
    row += 1
    for i, tonaj in enumerate(meta.get("rulolar_ton") or [], start=1):
        ws.cell(row=row, column=1, value=f"Rulo {i}")
        ws.cell(row=row, column=2, value=tonaj)
        row += 1
    row += 1

    siparisler = list(meta.get("siparisler") or [])
    if siparisler:
        _bolum_baslik(ws, row, "SİPARİŞ LİSTESİ")
        row += 1
        s1 = ws.cell(row=row, column=1, value="Sipariş No")
        s1.font = BASLIK_FONT
        s1.fill = ALT_BASLIK_DOLGU
        s2 = ws.cell(row=row, column=2, value="Miktar (m²)")
        s2.font = BASLIK_FONT
        s2.fill = ALT_BASLIK_DOLGU
        s3 = ws.cell(row=row, column=3, value="Panel Genişliği (m)")
        s3.font = BASLIK_FONT
        s3.fill = ALT_BASLIK_DOLGU
        s4 = ws.cell(row=row, column=4, value="Panel Uzunluğu (m)")
        s4.font = BASLIK_FONT
        s4.fill = ALT_BASLIK_DOLGU
        row += 1
        for j, s in enumerate(siparisler, start=1):
            ws.cell(row=row, column=1, value=f"Sipariş {j}")
            ws.cell(row=row, column=2, value=float(s.get("m2", 0) or 0))
            ws.cell(row=row, column=3, value=float(s.get("panelWidth", 0) or 0))
            ws.cell(row=row, column=4, value=float(s.get("panelLength", 1.0) or 1.0))
            row += 1
        row += 1

    maliyetler = meta.get("maliyetler") or {}
    if maliyetler:
        _bolum_baslik(ws, row, "MALİYET KATSAYILARI")
        row += 1
        if "fire_cost" in maliyetler:
            _yaz_etiket_deger(ws, row, "Fire Maliyeti (cf)", maliyetler["fire_cost"])
            row += 1
        if "setup_cost" in maliyetler:
            _yaz_etiket_deger(ws, row, "Setup Maliyeti - Rulo Açma (A)", maliyetler["setup_cost"])
            row += 1
        if "stock_cost" in maliyetler:
            _yaz_etiket_deger(ws, row, "Stok Tutma Maliyeti (h)", maliyetler["stock_cost"])
            row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20


def _sayfa_2_ozet(wb: Workbook, sonuc: Dict[str, Any], maliyetler: Dict[str, Any]) -> None:
    """
    Sayfa 2: Özet. Toplam maliyet, fire, stok, açılan rulo + maliyet kırılımı.

    Args:
        wb: Workbook
        sonuc: test_calistir veya solve_optimization özetli sözlük (toplam_fire, toplam_stok,
               toplam_maliyet, kullanilan_rulo_sayisi, phase, solver_status, failure_code, hints)
        maliyetler: fire_cost, setup_cost, stock_cost
    """
    ws = wb.create_sheet("Ozet")
    ws["A1"] = "KESME STOKU OPTİMİZASYON RAPORU - ÖZET"
    ws["A1"].font = BUYUK_BASLIK
    ws.merge_cells("A1:B1")

    toplam_fire = float(sonuc.get("toplam_fire") or 0.0)
    toplam_stok = float(sonuc.get("toplam_stok") or 0.0)
    toplam_elde = float(sonuc.get("toplam_elde_ton") or 0.0)
    acilan_rulo = int(sonuc.get("kullanilan_rulo_sayisi") or 0)
    fire_cost = float(maliyetler.get("fire_cost") or 0.0)
    stock_cost = float(maliyetler.get("stock_cost") or 0.0)
    setup_cost = float(maliyetler.get("setup_cost") or 0.0)
    toplam_fire_maliyet = float(sonuc.get("cost_fire_lira") or (toplam_fire * fire_cost))
    toplam_stok_maliyet = float(
        sonuc.get("cost_stock_lira") or ((toplam_stok + toplam_elde) * stock_cost),
    )
    stok_uretim_tl = float(sonuc.get("cost_stock_production_lira") or (toplam_stok * stock_cost))
    stok_elde_tl = float(sonuc.get("cost_stock_shelf_lira") or (toplam_elde * stock_cost))
    toplam_setup_maliyet = float(sonuc.get("cost_setup_lira") or (acilan_rulo * setup_cost))
    sira_ceza_tl = float(sonuc.get("cost_sequence_penalty_lira") or 0.0)
    toplam_maliyet = sonuc.get("toplam_maliyet")
    if toplam_maliyet is None:
        toplam_maliyet = toplam_fire_maliyet + toplam_stok_maliyet + toplam_setup_maliyet + sira_ceza_tl

    ozet_rows: List[Sequence[Any]] = [
        ["Çözüm Fazı", sonuc.get("phase") or ""],
        ["Çözücü Durumu", sonuc.get("solver_status") or ""],
        ["Hata / Başarısızlık Kodu", sonuc.get("failure_code") or ""],
        ["Toplam Maliyet (TL)", f"{float(toplam_maliyet):.2f}"],
        ["Toplam Fire (Ton)", f"{toplam_fire:.4f}"],
        ["Toplam Üretim Stoku (Ton)", f"{toplam_stok:.4f}"],
        ["Toplam Rafta Elde (Ton)", f"{toplam_elde:.4f}"],
        ["Açılan Rulo Sayısı", acilan_rulo],
        ["— Maliyet kalemleri (TL) —", ""],
        ["Fire (cf × fire tonu)", f"{toplam_fire_maliyet:.2f}"],
        ["Stok tutma — üretim stoku (h × ton)", f"{stok_uretim_tl:.2f}"],
        ["Stok tutma — rafta elde (h × ton)", f"{stok_elde_tl:.2f}"],
        ["Stok tutma toplamı", f"{toplam_stok_maliyet:.2f}"],
        ["Rulo açılış / kurulum (A × açılan rulo)", f"{toplam_setup_maliyet:.2f}"],
        ["Sıra cezası (varsa)", f"{sira_ceza_tl:.2f}"],
    ]
    rulo_degisim = sonuc.get("rulo_degisim_sayisi")
    if rulo_degisim is not None:
        ozet_rows.append(["Kesim Planı Rulo Değişim Sayısı", rulo_degisim])
    hat_gecis = sonuc.get("uretim_hatti_rulo_gecis_sayisi")
    if hat_gecis is not None:
        ozet_rows.append(["Üretim Hattı Rulo Geçiş Sayısı", hat_gecis])

    _tablo_yaz(ws, 3, ["Metrik", "Değer"], ozet_rows)

    hints = sonuc.get("hints") or []
    if hints:
        row = 4 + len(ozet_rows) + 1
        _bolum_baslik(ws, row, "İPUÇLARI / UYARILAR")
        row += 1
        for h in hints:
            ws.cell(row=row, column=1, value="-").font = BASLIK_FONT
            c = ws.cell(row=row, column=2, value=str(h))
            c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50


def _sayfa_4_rulo_durumu(wb: Workbook, results: Optional[Dict[str, Any]]) -> None:
    """
    Sayfa 3: Rulo Durumu. Rulo başına kapasite, kullanılan, kalan, stok, fire, durum.

    Args:
        wb: Workbook
        results: solve_optimization çıktısı (rollStatus)
    """
    ws = wb.create_sheet("Rulo_Durumu")
    ws["A1"] = "RULO DURUMU"
    ws["A1"].font = BUYUK_BASLIK
    ws.merge_cells("A1:I1")

    basliklar = [
        "Rulo ID",
        "Başlangıç Kapasitesi (Ton)",
        "Kullanılan (Ton)",
        "Kalan (Ton)",
        "Stok (Ton)",
        "Fire (Ton)",
        "Elde (Ton)",
        "Kullanılan Sipariş Sayısı",
        "Durum",
    ]
    satirlar: List[Sequence[Any]] = []
    if results and results.get("rollStatus"):
        for item in results.get("rollStatus") or []:
            toplam = float(item.get("totalTonnage", 0) or 0)
            kullanilan = float(item.get("used", 0) or 0)
            kalan = toplam - kullanilan
            stok = float(item.get("stock", 0) or 0)
            fire = float(item.get("fire", 0) or 0)
            elde = float(item.get("unusedRollTonnage", 0) or 0)
            orders_used = int(item.get("ordersUsed", 0) or 0)
            if kullanilan < 0.0001:
                durum = "Kullanılmadı"
                stok_g, fire_g = "0.0000", "0.0000"
            elif fire > 0.0001:
                durum = "Fire"
                stok_g, fire_g = f"{stok:.4f}", f"{fire:.4f}"
            elif stok > 0.0001:
                durum = "Stok"
                stok_g, fire_g = f"{stok:.4f}", f"{fire:.4f}"
            else:
                durum = "Tamamen Kullanıldı"
                stok_g, fire_g = f"{stok:.4f}", f"{fire:.4f}"
            elde_g = f"{elde:.4f}"
            satirlar.append(
                [
                    f"Rulo {item.get('rollId', '')}",
                    f"{toplam:.2f}",
                    f"{kullanilan:.4f}",
                    f"{kalan:.4f}",
                    stok_g,
                    fire_g,
                    elde_g,
                    orders_used,
                    durum,
                ]
            )
    else:
        satirlar.append(["Çözüm yok", "", "", "", "", "", "", "", ""])

    _tablo_yaz(ws, 3, basliklar, satirlar)

    for i, w in enumerate([15, 25, 18, 18, 15, 15, 15, 22, 22], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sayfa_5_detayli_kesim(wb: Workbook, results: Optional[Dict[str, Any]]) -> None:
    """
    Sayfa 4: Detaylı Kesim Planı. Rulo başına tek satırda kesilen parçalar.

    Args:
        wb: Workbook
        results: solve_optimization çıktısı
    """
    ws = wb.create_sheet("Detayli_Kesim_Plani")
    ws["A1"] = "DETAYLI KESİM PLANI"
    ws["A1"].font = BUYUK_BASLIK
    ws.merge_cells("A1:F1")

    basliklar = [
        "Rulo No",
        "Başlangıç Kapasitesi (Ton)",
        "Kullanılan Miktar (Ton)",
        "Kalan (Fire veya Stok) (Ton)",
        "Durum",
        "Kesilen Parçalar ve Eşleşen Siparişler",
    ]
    for col_idx, baslik in enumerate(basliklar, start=1):
        c = ws.cell(row=3, column=col_idx, value=baslik)
        c.font = BASLIK_FONT
        c.fill = BASLIK_DOLGU
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if not results or not results.get("rollStatus"):
        ws.cell(row=4, column=1, value="Çözüm yok (ön kontrol veya Infeasible)")
        for i, w in enumerate([12, 22, 20, 22, 18, 40], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return

    by_roll: Dict[int, List[Dict[str, Any]]] = {}
    for item in results.get("cuttingPlan") or []:
        rid = int(item.get("rollId", 0))
        by_roll.setdefault(rid, []).append(item)

    row_idx = 4
    for item in results.get("rollStatus") or []:
        rid = int(item.get("rollId", 0))
        toplam = float(item.get("totalTonnage", 0) or 0)
        kullanilan = float(item.get("used", 0) or 0)
        fire = float(item.get("fire", 0) or 0)
        stok = float(item.get("stock", 0) or 0)
        if kullanilan < 0.0001:
            durum, kalan_miktar = "Kullanılmadı", toplam
        elif fire > 0.0001:
            durum, kalan_miktar = "Fire Var", fire
        elif stok > 0.0001:
            durum, kalan_miktar = "Stok Oluştu", stok
        else:
            durum, kalan_miktar = "Tam Kullanıldı", 0.0

        parcalar_list: List[str] = []
        for c in by_roll.get(rid, []):
            parcalar_list.append(
                f"- {float(c.get('tonnage', 0) or 0):.4f} ton "
                f"({float(c.get('m2', 0) or 0):.2f} m²) Sipariş {c.get('orderId', '')}"
            )
        parcalar = "\n".join(parcalar_list) if parcalar_list else "Kesim yapılmadı"

        ws.cell(row=row_idx, column=1, value=f"Rulo {rid}")
        ws.cell(row=row_idx, column=2, value=f"{toplam:.2f}")
        ws.cell(row=row_idx, column=3, value=f"{kullanilan:.4f}")
        ws.cell(row=row_idx, column=4, value=f"{kalan_miktar:.4f}")
        ws.cell(row=row_idx, column=5, value=durum)
        ws.cell(row=row_idx, column=6, value=parcalar)
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        n = max(1, parcalar.count("\n") + 1)
        ws.row_dimensions[row_idx].height = min(15 * n, 120)
        row_idx += 1

    for i, w in enumerate([12, 22, 22, 22, 18, 50], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sayfa_6_uretim_adimlari(wb: Workbook, results: Optional[Dict[str, Any]]) -> None:
    """
    Sayfa 5: Üretim Adımları. Çift yüzey eş zamanlı üretim adımlarını (lineSchedule)
    satır satır gösterir: her adımda hangi üst + alt rulo, hangi siparişi kaç ton kesiyor.

    Örnek satır mantığı:
      Adım 1 · Sipariş 3 · Üst Rulo 1 (4.00 t) · Alt Rulo 5 (4.00 t) · aksiyon: üretim başladı
      Adım 2 · Sipariş 3 · Üst Rulo 1 (2.00 t) · Alt Rulo 7 (2.00 t) · Alt: R7 takıldı
      Adım 3 · Sipariş 1 · Üst Rulo 2 (2.00 t) · Alt Rulo 6 (2.00 t) · Üst+Alt: rulo değişimi

    Args:
        wb: Workbook
        results: solve_optimization çıktısı (lineSchedule, lineTransitionsSummary)
    """
    ws = wb.create_sheet("Uretim_Adimlari")
    ws["A1"] = "ÜRETİM ADIMLARI (Eş Zamanlı Üst + Alt Yüzey)"
    ws["A1"].font = BUYUK_BASLIK
    ws.merge_cells("A1:I1")

    if not results or not results.get("lineSchedule"):
        ws.cell(row=3, column=1, value="Çözüm yok — üretim adımları çıkarılamadı.")
        for col_idx, w in enumerate([8, 10, 12, 12, 12, 12, 15, 50], start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        return

    aciklama = (
        "Her adım, üst ve alt yüzeyin eş zamanlı çalıştığı tek bir üretim dilimini temsil eder. "
        "Üst + Alt rulo çiftleri aynı anda sipariş kesiminde kullanılır. 'Aksiyon' sütunu bu adımda "
        "rulo tak-çıkar ve sipariş değişimi olup olmadığını özetler."
    )
    c = ws.cell(row=2, column=1, value=aciklama)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:I2")

    basliklar = [
        "Adım",
        "Sipariş",
        "Üst Rulo",
        "Alt Rulo",
        "Üst Ton",
        "Alt Ton",
        "Üst Aksiyon",
        "Alt Aksiyon",
        "Aksiyon Özeti",
    ]
    satirlar: List[Sequence[Any]] = []
    for ev in results.get("lineSchedule") or []:
        upper_ton = 0.0
        lower_ton = 0.0
        for c_ in ev.get("cuts") or []:
            rid = c_.get("rollId")
            if rid is None:
                continue
            if int(rid) == int(ev.get("upperRollId") or -1):
                upper_ton += float(c_.get("tonnage") or 0.0)
            if int(rid) == int(ev.get("lowerRollId") or -1):
                lower_ton += float(c_.get("tonnage") or 0.0)
        satirlar.append(
            [
                int(ev.get("step") or 0),
                f"Sipariş {ev.get('orderId', '')}",
                f"Rulo {ev.get('upperRollId', '-')}" if ev.get("upperRollId") is not None else "—",
                f"Rulo {ev.get('lowerRollId', '-')}" if ev.get("lowerRollId") is not None else "—",
                f"{upper_ton:.4f}" if upper_ton > 0 else "0.0000",
                f"{lower_ton:.4f}" if lower_ton > 0 else "0.0000",
                ev.get("upperAction", ""),
                ev.get("lowerAction", ""),
                ev.get("actionSummary", ""),
            ]
        )

    _tablo_yaz(ws, 4, basliklar, satirlar)

    for col_idx, w in enumerate([8, 12, 12, 12, 12, 12, 14, 14, 55], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    lts = results.get("lineTransitionsSummary") or {}
    sonuc_satir = 5 + len(satirlar) + 1
    ws.cell(row=sonuc_satir, column=1, value="ÖZET").font = ORTA_BASLIK
    ws.cell(row=sonuc_satir + 1, column=1, value="Toplam adım sayısı")
    ws.cell(row=sonuc_satir + 1, column=2, value=int(lts.get("stepCount") or len(satirlar)))
    ws.cell(row=sonuc_satir + 2, column=1, value="Toplam rulo tak-çıkar (totalChanges)")
    ws.cell(row=sonuc_satir + 2, column=2, value=int(lts.get("totalChanges") or 0))
    ws.cell(row=sonuc_satir + 3, column=1, value="Eş zamanlı üst+alt değişim")
    ws.cell(row=sonuc_satir + 3, column=2, value=int(lts.get("synchronousChanges") or 0))
    ws.cell(row=sonuc_satir + 4, column=1, value="Bağımsız (yalnız üst veya yalnız alt) değişim")
    ws.cell(row=sonuc_satir + 4, column=2, value=int(lts.get("independentChanges") or 0))


def _sayfa_grafikler(wb: Workbook, grafik_yollari: Sequence[str]) -> None:
    """
    Grafikler sayfası. Verilen PNG dosyalarını sırayla dikey ekler.

    Args:
        wb: Workbook
        grafik_yollari: Embed edilecek PNG dosyalarının mutlak yolları
    """
    if not grafik_yollari:
        return
    ws = wb.create_sheet("Grafikler")
    ws["A1"] = "GRAFİKLER"
    ws["A1"].font = BUYUK_BASLIK
    ws.merge_cells("A1:H1")

    satir = 3
    for p in grafik_yollari:
        if not p or not os.path.exists(p):
            continue
        ws.cell(row=satir, column=1, value=os.path.basename(p)).font = BASLIK_FONT
        satir += 1
        img = XlImage(p)
        img.width = 720
        img.height = 432
        ws.add_image(img, f"A{satir}")
        satir += 24


def build_cozum_raporu_xlsx(
    scenario_meta: Dict[str, Any],
    sonuc: Dict[str, Any],
    output_path: str,
    *,
    grafik_yollari: Optional[Iterable[str]] = None,
) -> str:
    """
    main/sonuclar/cozum_raporu_*.xlsx ile uyumlu (Kesim_Plani hariç) + opsiyonel Grafikler
    sayfası üretir.

    Args:
        scenario_meta: Senaryo meta sözlüğü (senaryo_adi, guvenlik_payi, kalinlik_mm,
            yogunluk_g_cm3, rulolar_ton, siparisler, maliyetler, toplam_talep_ton,
            toplam_rulo_kapasitesi_ton, max_siparis_per_rulo, max_rulo_per_siparis, aciklama)
        sonuc: test_calistir veya solve_optimization çıktısından özetlenmiş sözlük
            (toplam_fire, toplam_stok, toplam_maliyet, kullanilan_rulo_sayisi,
            rulo_degisim_sayisi, uretim_hatti_rulo_gecis_sayisi, phase, solver_status,
            failure_code, hints, raw["results"])
        output_path: Yazılacak .xlsx yolu (dizin yoksa oluşturulur)
        grafik_yollari: Grafikler sayfasına gömülecek PNG dosyaları (opsiyonel)

    Returns:
        Yazılan dosyanın tam yolu
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    results = None
    raw = sonuc.get("raw")
    if isinstance(raw, dict):
        results = raw.get("results")

    maliyetler = scenario_meta.get("maliyetler") or {}
    _sayfa_1_veri_girisi(wb, scenario_meta)
    _sayfa_2_ozet(wb, sonuc, maliyetler)
    _sayfa_4_rulo_durumu(wb, results)
    _sayfa_5_detayli_kesim(wb, results)
    _sayfa_6_uretim_adimlari(wb, results)

    if grafik_yollari:
        _sayfa_grafikler(wb, [p for p in grafik_yollari if p])

    wb.save(output_path)
    return output_path


def scenario_meta_from_test_calistir(
    *,
    senaryo_adi: str,
    sonuc: Dict[str, Any],
    siparisler: Sequence[Dict[str, Any]],
    kalinlik_mm: float = 0.75,
    yogunluk_g_cm3: float = 7.85,
    fire_cost: float = 100.0,
    setup_cost: float = 50.0,
    stock_cost: float = 30.0,
    max_siparis_per_rulo: Optional[int] = None,
    max_rulo_per_siparis: Optional[int] = None,
    guvenlik_payi: Optional[float] = None,
    aciklama: Optional[str] = None,
) -> Dict[str, Any]:
    """
    test_calistir çıktısı + girdi parametrelerinden scenario_meta sözlüğü üretir.

    Args:
        senaryo_adi: Senaryo başlığı
        sonuc: test_calistir dönüşü
        siparisler: Girdi sipariş listesi (m2, panelWidth, panelLength)
        kalinlik_mm: Malzeme kalınlığı
        yogunluk_g_cm3: Malzeme yoğunluğu
        fire_cost: Fire birim maliyeti
        setup_cost: Kurulum birim maliyeti
        stock_cost: Stok birim maliyeti
        max_siparis_per_rulo: Rulo başına max sipariş
        max_rulo_per_siparis: Sipariş başına max rulo
        guvenlik_payi: Varsa %
        aciklama: Kısa açıklama

    Returns:
        scenario_meta sözlüğü
    """
    rulolar_str = str(sonuc.get("rulo_kapasiteleri_str") or "")
    rulolar_ton: List[float] = []
    if rulolar_str:
        try:
            rulolar_ton = [float(x) for x in rulolar_str.split("+") if x.strip()]
        except ValueError:
            rulolar_ton = []
    return {
        "senaryo_adi": senaryo_adi,
        "guvenlik_payi": guvenlik_payi,
        "aciklama": aciklama,
        "kalinlik_mm": kalinlik_mm,
        "yogunluk_g_cm3": yogunluk_g_cm3,
        "rulolar_ton": rulolar_ton,
        "siparisler": [dict(o) for o in siparisler],
        "maliyetler": {
            "fire_cost": fire_cost,
            "setup_cost": setup_cost,
            "stock_cost": stock_cost,
        },
        "toplam_talep_ton": sonuc.get("toplam_talep_ton") or 0.0,
        "toplam_rulo_kapasitesi_ton": sonuc.get("toplam_rulo_kapasitesi_ton") or 0.0,
        "max_siparis_per_rulo": max_siparis_per_rulo,
        "max_rulo_per_siparis": max_rulo_per_siparis,
    }


def sonuc_from_optimizer_results(results: Dict[str, Any]) -> Dict[str, Any]:
    """
    solve_optimization çıktısından build_cozum_raporu_xlsx ile uyumlu özet sözlüğü üretir.

    Args:
        results: Kesim planı, özet ve üretim hattı alanlarını içeren çözücü sonucu

    Returns:
        test_calistir biçimine yakın `sonuc` sözlüğü (raw.results ile tam detay)
    """
    sm = results.get("summary") or {}
    opened = int(sm.get("openedRolls", 0) or 0)
    tf = float(sm.get("totalFire", 0.0) or 0.0)
    ts = float(sm.get("totalStock", 0.0) or 0.0)
    tc = float(sm.get("totalCost", 0.0) or 0.0)
    hints: List[str] = []
    seq_pen = float(sm.get("sequencePenalty", 0) or 0)
    if seq_pen > 1e-9:
        hints.append(f"Sıra cezası (siparişe dönüş): {seq_pen:.4f}")
    viol_n = int(sm.get("interleavingViolationCount", 0) or 0)
    if viol_n > 0:
        hints.append(f"Sıra ihlal sayısı: {viol_n}")

    tu = float(sm.get("totalUnusedInventoryTon", 0) or 0.0)
    sonuc: Dict[str, Any] = {
        "kullanilan_rulo_sayisi": opened,
        "toplam_fire": round(tf, 4),
        "toplam_stok": round(ts, 4),
        "toplam_elde_ton": round(tu, 4),
        "toplam_maliyet": round(tc, 2),
        "cost_fire_lira": round(float(sm.get("costFireLira", 0) or 0), 2),
        "cost_stock_lira": round(float(sm.get("costStockLira", 0) or 0), 2),
        "cost_stock_production_lira": round(float(sm.get("costStockProductionLira", 0) or 0), 2),
        "cost_stock_shelf_lira": round(float(sm.get("costStockShelfLira", 0) or 0), 2),
        "cost_setup_lira": round(float(sm.get("costSetupLira", 0) or 0), 2),
        "cost_sequence_penalty_lira": round(float(sm.get("costSequencePenaltyLira", 0) or 0), 2),
        "mesaj": "",
        "phase": "solver",
        "solver_status": "Optimal",
        "failure_code": None,
        "hints": hints,
        "raw": {"results": results},
    }
    rulo_deg = int(
        sm.get("rollChangeCount", results.get("rollChangeCount", 0)) or 0
    )
    sync_ihlal = int(
        sm.get("surfaceSyncViolations", results.get("surfaceSyncViolations", 0)) or 0
    )
    lts = results.get("lineTransitionsSummary") or {}
    sonuc["rulo_degisim_sayisi"] = rulo_deg
    sonuc["yuzey_es_zaman_ihlal_sayisi"] = sync_ihlal
    sonuc["uretim_hatti_rulo_gecis_sayisi"] = int(lts.get("totalChanges", 0) or 0)
    sonuc["uretim_hatti_es_zamanli_gecis_sayisi"] = int(
        lts.get("synchronousChanges", 0) or 0
    )
    sonuc["uretim_hatti_bagimsiz_gecis_sayisi"] = int(
        lts.get("independentChanges", 0) or 0
    )
    return sonuc


def scenario_meta_from_dashboard_inputs(
    *,
    senaryo_adi: str,
    kalinlik_mm: float,
    yogunluk_g_cm3: float,
    rulolar_ton: Sequence[float],
    siparisler: Sequence[Dict[str, Any]],
    fire_cost: float,
    setup_cost: float,
    stock_cost: float,
    toplam_talep_ton: float,
    toplam_rulo_kapasitesi_ton: float,
    guvenlik_payi: Optional[float] = None,
    max_siparis_per_rulo: Optional[int] = None,
    max_rulo_per_siparis: Optional[int] = None,
    aciklama: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Dashboard /api/optimize isteğinden scenario_meta üretir (Kullanici_Veri_Girisi sayfası).

    Args:
        senaryo_adi: Rapor başlığı (kullanıcı açıklaması veya varsayılan)
        kalinlik_mm: Malzeme kalınlığı
        yogunluk_g_cm3: Yoğunluk (g/cm³)
        rulolar_ton: Kullanılan rulo tonaj listesi
        siparisler: Sipariş satırları (m2, panelWidth, panelLength)
        fire_cost: Fire birim maliyeti
        setup_cost: Rulo açma maliyeti
        stock_cost: Stok birim maliyeti
        toplam_talep_ton: Çift yüzey dahil toplam talep (ton)
        toplam_rulo_kapasitesi_ton: Toplam rulo kapasitesi (ton)
        guvenlik_payi: Varsa güvenlik stoğu %
        max_siparis_per_rulo: Rulo başına max sipariş
        max_rulo_per_siparis: Sipariş başına max rulo
        aciklama: Ek not (senkron seviye, mod vb.)

    Returns:
        build_cozum_raporu_xlsx için scenario_meta sözlüğü
    """
    return {
        "senaryo_adi": senaryo_adi,
        "guvenlik_payi": guvenlik_payi,
        "aciklama": aciklama,
        "kalinlik_mm": kalinlik_mm,
        "yogunluk_g_cm3": yogunluk_g_cm3,
        "rulolar_ton": [float(x) for x in rulolar_ton],
        "siparisler": [dict(o) for o in siparisler],
        "maliyetler": {
            "fire_cost": fire_cost,
            "setup_cost": setup_cost,
            "stock_cost": stock_cost,
        },
        "toplam_talep_ton": float(toplam_talep_ton),
        "toplam_rulo_kapasitesi_ton": float(toplam_rulo_kapasitesi_ton),
        "max_siparis_per_rulo": max_siparis_per_rulo,
        "max_rulo_per_siparis": max_rulo_per_siparis,
    }


def karsilastirma_xlsx(
    rows: Sequence[Dict[str, Any]],
    output_path: str,
    *,
    grafik_yollari: Optional[Iterable[str]] = None,
    baslik: str = "SENARYO KARŞILAŞTIRMA",
    ek_sheetler: Optional[Dict[str, Sequence[Sequence[Any]]]] = None,
) -> str:
    """
    Birden fazla senaryo/veri satırını karşılaştırma XLSX'ine yazar.

    Args:
        rows: Her öğe "Senaryo", "Toplam Fire (Ton)", "Toplam Stok (Ton)", vb. içeren sözlük
        output_path: Dosya yolu
        grafik_yollari: Grafikler sayfasına gömülecek PNG yolları
        baslik: Ozet sayfasındaki başlık
        ek_sheetler: Ek tablolar ({sheet_adi: [[header..], [row..], ...]})

    Returns:
        Yazılan dosyanın yolu
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Ozet")
    ws["A1"] = baslik
    ws["A1"].font = BUYUK_BASLIK

    if rows:
        kolonlar = list(rows[0].keys())
        satirlar = [[r.get(k, "") for k in kolonlar] for r in rows]
        _tablo_yaz(ws, 3, kolonlar, satirlar)
        for idx, _ in enumerate(kolonlar, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = 22

    if ek_sheetler:
        for sheet_adi, data in ek_sheetler.items():
            wsx = wb.create_sheet(sheet_adi[:31])
            if data:
                basliklar = list(data[0])
                veri_satirlari = [list(r) for r in data[1:]]
                _tablo_yaz(wsx, 1, basliklar, veri_satirlari)
                for idx, _ in enumerate(basliklar, start=1):
                    wsx.column_dimensions[get_column_letter(idx)].width = 22

    if grafik_yollari:
        _sayfa_grafikler(wb, [p for p in grafik_yollari if p])

    wb.save(output_path)
    return output_path
